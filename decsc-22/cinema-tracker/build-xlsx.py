#!/usr/bin/env python3
"""Build the complete cinema-tracker.xlsx with all 5 tabs, formulas, dropdowns,
and conditional formatting.

Output: cinema-tracker/cinema-tracker.xlsx (drag-and-drop to Google Drive — it
auto-converts to a Google Sheet preserving everything except Apps Script, which
is added separately).
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.chart import PieChart, BarChart, Reference

ROOT = Path(__file__).parent
DATA = ROOT / "data"
OUT = ROOT / "cinema-tracker.xlsx"

# ─── Load cinema data ────────────────────────────────────────────────
cinemas = json.loads((DATA / "cinema-coordinates.json").read_text())
geocoded = sorted(
    [c for c in cinemas if c.get("lat")],
    key=lambda c: (c["platform"], c["name"]),
)

# ─── Workbook ────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
SUBHEADER_FONT = Font(bold=True)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

# ─── Sheet: Cinemas (reference data) ─────────────────────────────────
ws = wb.create_sheet("Cinemas")
headers = ["Platform", "CinemaId", "Name", "Address", "City", "Lat", "Lng"]
ws.append(headers)
style_header(ws, 1, len(headers))
for c in geocoded:
    ws.append([
        c["platform"], c["id"], c["name"],
        c.get("address", ""), c.get("city", ""),
        c["lat"], c["lng"],
    ])
ws.freeze_panes = "A2"
widths = [12, 12, 40, 50, 20, 11, 11]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ─── Sheet: Settings ─────────────────────────────────────────────────
ws = wb.create_sheet("Settings", 0)
ws["A1"] = "🎬 PH Cinema Price Tracker — Settings"
ws["A1"].font = Font(size=16, bold=True, color="1F4E79")
ws.merge_cells("A1:C1")
ws.row_dimensions[1].height = 30

ws["A3"] = "Setting"; ws["B3"] = "Value"; ws["C3"] = "Notes"
style_header(ws, 3, 3)

settings_rows = [
    ("Your Location (preset)", "Glorietta", "Pick from dropdown — drives distance calc"),
    ("Your Latitude",  "=IFERROR(VLOOKUP($B$4, Cinemas!C:G, 4, FALSE), 14.5547)",
        "Auto-fills from preset, OR overwrite manually"),
    ("Your Longitude", "=IFERROR(VLOOKUP($B$4, Cinemas!C:G, 5, FALSE), 121.0244)",
        "Auto-fills from preset, OR overwrite manually"),
    ("Max Distance (km)", 20, "Only show cinemas within this radius"),
    ("Max Budget (PHP)",  450, "Hide showtimes above this price"),
    ("Movie Filter (optional)", "", "Type a movie name to filter (case-insensitive partial match) — leave blank for all"),
]
for i, (k, v, note) in enumerate(settings_rows, start=4):
    ws.cell(row=i, column=1, value=k).font = SUBHEADER_FONT
    ws.cell(row=i, column=2, value=v)
    ws.cell(row=i, column=3, value=note).alignment = Alignment(wrap_text=True)

# Dropdown for B4 (location preset)
dv = DataValidation(
    type="list",
    formula1=f"=Cinemas!$C$2:$C${len(geocoded) + 1}",
    allow_blank=False,
)
dv.error = "Pick a cinema name from the list"
dv.errorTitle = "Invalid location"
ws.add_data_validation(dv)
dv.add("B4")

# Highlight the editable cells
for r in [4, 7, 8, 9]:
    ws.cell(row=r, column=2).fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 60

ws["A12"] = "How to use:"
ws["A12"].font = Font(bold=True, color="1F4E79")
help_text = [
    "1. Pick your nearest cinema as 'Your Location' (B4 dropdown)",
    "2. Set Max Distance + Max Budget below",
    "3. Click 🎬 Cinema Tracker → Refresh Showtimes (in the menu bar)",
    "4. Open the Dashboard tab to see ranked results",
    "5. Track your movie spending in the Budget tab",
]
for i, line in enumerate(help_text, start=13):
    ws.cell(row=i, column=1, value=line)
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

# ─── Sheet: Showtimes (placeholder; Apps Script will fill) ───────────
ws = wb.create_sheet("Showtimes")
headers = ["Platform", "CinemaId", "Cinema", "Movie", "Date", "Time",
           "Price (PHP)", "Genre", "Venue", "Link"]
ws.append(headers)
style_header(ws, 1, len(headers))
ws["L1"] = "Last refreshed:"
ws["M1"] = "(click Refresh in 🎬 menu)"
ws["L1"].font = Font(italic=True, color="808080")
ws["M1"].font = Font(italic=True, color="808080")

ws["A2"] = "(Click 🎬 Cinema Tracker → Refresh Showtimes to load live data)"
ws["A2"].font = Font(italic=True, color="808080")
ws.merge_cells("A2:J2")

ws.freeze_panes = "A2"
widths = [12, 14, 36, 36, 12, 8, 12, 14, 25, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ─── Sheet: Dashboard (live calc) ────────────────────────────────────
ws = wb.create_sheet("Dashboard")
ws["A1"] = "🎬 Dashboard — Cheapest movies near you"
ws["A1"].font = Font(size=16, bold=True, color="1F4E79")
ws.merge_cells("A1:I1")
ws.row_dimensions[1].height = 30

ws["A3"] = ("This list updates live from your Settings tab. "
            "Refresh data via 🎬 Cinema Tracker → Refresh Showtimes.")
ws["A3"].font = Font(italic=True, color="808080")
ws.merge_cells("A3:I3")

dash_headers = ["Platform", "Cinema", "Movie", "Date", "Time",
                "Price (PHP)", "Distance (km)", "Link"]
for i, h in enumerate(dash_headers, 1):
    ws.cell(row=5, column=i, value=h)
style_header(ws, 5, len(dash_headers))

# Dashboard rows — each row uses a single ARRAYFORMULA-friendly approach:
# We use individual SORTN over Showtimes joined with a distance column.
# Since openpyxl doesn't support QUERY, we use a more compatible approach:
# - Build helper columns K-Q on Showtimes for distance
# - Sort + filter in Dashboard via FILTER + SORT + LARGE/SMALL
#
# For Google Sheets compatibility, we use this simpler formula in A6:
# =SORT(FILTER(...), price ASC, distance ASC)
#
# But because xlsx → Sheets converts dynamic arrays well, we use FILTER+SORT.
ws["A6"] = (
    '=IFERROR('
    'SORT('
    'FILTER('
    '{Showtimes!A2:A1000, Showtimes!C2:C1000, Showtimes!D2:D1000, '
    'Showtimes!E2:E1000, Showtimes!F2:F1000, Showtimes!G2:G1000, '
    'ARRAYFORMULA(IF(Showtimes!A2:A1000="", "", '
    '6371*2*ASIN(SQRT('
    'SIN((RADIANS(IFERROR(VLOOKUP(Showtimes!B2:B1000, Cinemas!B:G, 5, FALSE), Settings!$B$5))-RADIANS(Settings!$B$5))/2)^2'
    '+COS(RADIANS(Settings!$B$5))*COS(RADIANS(IFERROR(VLOOKUP(Showtimes!B2:B1000, Cinemas!B:G, 5, FALSE), Settings!$B$5)))'
    '*SIN((RADIANS(IFERROR(VLOOKUP(Showtimes!B2:B1000, Cinemas!B:G, 6, FALSE), Settings!$B$6))-RADIANS(Settings!$B$6))/2)^2'
    ')))), '
    'Showtimes!J2:J1000}, '
    'Showtimes!A2:A1000<>"", '
    'Showtimes!G2:G1000>0, '
    'Showtimes!G2:G1000<=Settings!$B$8, '
    'IF(Settings!$B$9="", TRUE, '
    'IFERROR(SEARCH(LOWER(Settings!$B$9), LOWER(Showtimes!D2:D1000))>0, FALSE))'
    '), '
    '6, TRUE, 7, TRUE), '
    '"No showings match your filters yet — refresh data and check your Settings."'
    ')'
)

ws.freeze_panes = "A6"
widths = [12, 36, 36, 12, 8, 12, 14, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Conditional formatting on Dashboard
green_rule = CellIsRule(operator="lessThan", formula=["350"], fill=GREEN_FILL)
ws.conditional_formatting.add("F6:F1000", green_rule)
orange_rule = CellIsRule(operator="between", formula=["10", "20"], fill=ORANGE_FILL)
red_rule = CellIsRule(operator="greaterThan", formula=["20"], fill=RED_FILL)
ws.conditional_formatting.add("G6:G1000", orange_rule)
ws.conditional_formatting.add("G6:G1000", red_rule)

# ─── Sheet: Budget ───────────────────────────────────────────────────
ws = wb.create_sheet("Budget")
ws["A1"] = "💸 My Movie Budget Tracker"
ws["A1"].font = Font(size=16, bold=True, color="1F4E79")
ws.merge_cells("A1:G1")
ws.row_dimensions[1].height = 30

bud_headers = ["Date", "Movie", "Cinema", "Tickets",
               "Price per Ticket (PHP)", "Total (PHP)", "Notes"]
for i, h in enumerate(bud_headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(bud_headers))

# Sample rows
sample_rows = [
    ("2026-04-15", "Lee Cronin's The Mummy", "Robinsons Manila", 2, 330, None, "Friday date night"),
    ("2026-04-22", "Almost Us", "SM City North Edsa", 1, 420, None, "Solo"),
    ("2026-05-02", "The Devil Wears Prada 2", "Greenbelt 5", 3, 500, None, "With group"),
]
for i, row in enumerate(sample_rows, start=4):
    for j, val in enumerate(row, start=1):
        if j == 6:  # Total auto-formula
            ws.cell(row=i, column=j, value=f"=IF(A{i}=\"\",\"\",D{i}*E{i})")
        else:
            ws.cell(row=i, column=j, value=val)

# Pre-fill total formula in rows 7-100 (so user can just type in A:E and it works)
for i in range(7, 101):
    ws.cell(row=i, column=6, value=f"=IF(A{i}=\"\",\"\",D{i}*E{i})")

# Currency format
for col in [5, 6]:
    for i in range(4, 101):
        ws.cell(row=i, column=col).number_format = "\"₱\"#,##0.00"

# Summary block (right side)
ws["I3"] = "Summary"
ws["I3"].font = Font(bold=True, size=14, color="1F4E79")
summary = [
    ("Total spent YTD",       "=SUM(F4:F1000)",                                                         "\"₱\"#,##0.00"),
    ("Tickets bought",         "=SUM(D4:D1000)",                                                         "0"),
    ("Avg price / ticket",     "=IFERROR(I5/I6,0)",                                                      "\"₱\"#,##0.00"),
    ("Most-visited cinema",    '=IFERROR(INDEX(C4:C1000,MATCH(MAX(COUNTIF(C4:C1000,C4:C1000)),COUNTIF(C4:C1000,C4:C1000),0)),"-")', None),
    ("This month spent",       "=SUMIFS(F:F,A:A,\">=\"&EOMONTH(TODAY(),-1)+1,A:A,\"<=\"&EOMONTH(TODAY(),0))", "\"₱\"#,##0.00"),
    ("Movies this year",       "=COUNTIFS(A4:A1000,\">=\"&DATE(YEAR(TODAY()),1,1),A4:A1000,\"<=\"&DATE(YEAR(TODAY()),12,31))", "0"),
]
for i, (label, formula, fmt) in enumerate(summary, start=5):
    ws.cell(row=i, column=8, value=label).font = SUBHEADER_FONT
    cell = ws.cell(row=i, column=9, value=formula)
    if fmt:
        cell.number_format = fmt

ws.freeze_panes = "A4"
widths = [12, 32, 32, 9, 18, 13, 25, 4, 22, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Pie chart: spend by cinema
pie = PieChart()
pie.title = "Spend by cinema"
data = Reference(ws, min_col=6, min_row=3, max_row=20)
cats = Reference(ws, min_col=3, min_row=4, max_row=20)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)
ws.add_chart(pie, "I12")

# ─── Apps Script note ────────────────────────────────────────────────
# (Apps Script can't be embedded in xlsx — must be added in Google Sheets after upload)

# Reorder sheets: Settings first, then Dashboard, then helpers
wb._sheets = [
    wb["Settings"],
    wb["Dashboard"],
    wb["Showtimes"],
    wb["Budget"],
    wb["Cinemas"],
]
wb.active = 0

wb.save(OUT)
print(f"✓ Built {OUT}")
print(f"  Cinemas tab: {len(geocoded)} rows of geocoded venues")
print(f"  File size: {OUT.stat().st_size:,} bytes")
